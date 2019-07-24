#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jul 22 14:04:52 2019

@author: rosechrin
"""
import datetime 
from datetime import date
import math
import pandas as pd
import openpyxl

#This class creates an instance of a Bond based on a CUISP nummber 
class Bond:
    def __init__ (self, settlement_date, call_date, next_coupon_date, maturity_date, annual_coupon_rate, purchase_price, redemption, payment_frequency):
        self.settlement_date = settlement_date 
        self.call_date = call_date 
        self.next_coupon_date = next_coupon_date
        self.maturity_date = maturity_date 
        self.annual_coupon_rate = annual_coupon_rate
        self.purchase_price = purchase_price 
        self.redemption = redemption
        self.payment_frequency = payment_frequency 

#format for printing bond object        
    def __str__(self):
        bondInformation = str(self.settlement_date) + '\n' + str(self.call_date) + "\n" + \
                            str(self.maturity_date) + '\n' + str(self.annual_coupon_rate) + \
                            '\n' + str(self.purchase_price) + '\n' + str(self.redemption) + \
                            '\n' + str(self.payment_frequency) + '\n'
        return bondInformation
    
    def __repr__(self):
        return self.__str__()
    

#calculates the number of days to the maturity 
#date by using given call date and a settlement date two 
#days from now        
    def years_to_maturity(self):
        settlement_list = self.settlement_date.split("/",2)
        settlement = date(int(settlement_list[0]),int(settlement_list[1]), int(settlement_list[2]))
        if self.maturity_date != "N.A.":
            maturity_list = self.maturity_date.split("/",2)
            maturity = date(int(maturity_list[2]),int(maturity_list[0]),int(maturity_list[1]))
            days_to_mature = (maturity - settlement).days 
        else:
            maturity = date(2099,12,31)
            days_to_mature = (maturity - settlement).days 
        return (int(days_to_mature/365))

#calculates the number of days to the projected call 
#date by using given call date and a settlement date two 
#days from now         
    def years_to_call(self):
      settlement_list = self.settlement_date.split("/",2)
      settlement = date(int(settlement_list[0]),int(settlement_list[1]), int(settlement_list[2]))
      maturity_list = self.call_date.split("/",2)
      maturity = date(int(maturity_list[2]),int(maturity_list[0]),int(maturity_list[1]))
      days_to_mature = (maturity - settlement).days 
      if days_to_mature < 0:
          maturity_list = self.next_coupon_date.split("/",2)
          maturity = date(int(maturity_list[2]),int(maturity_list[0]),int(maturity_list[1]))
          days_to_mature = (maturity - settlement).days
      return int(days_to_mature/365)


#calculates the present value of the security using the 
#yield formula and number of years to the maturity date     
    def present_value(self,coupon,rate):
        face_value = self.redemption
        periods = self.years_to_maturity() * self.payment_frequency
        total_value = 0
        for n in range(1, periods + 1):
            total_value += coupon / math.pow((1+rate),n)           
        total_value += face_value / math.pow((1+rate),periods)
        return total_value

#calculates the present value of the security using the 
#yield formula and number of years to call date 
    def present_value2(self,coupon,rate):
        face_value = self.redemption
        periods = self.years_to_call() * self.payment_frequency
        total_value = 0
        for n in range(1, periods + 1):
            total_value += coupon / math.pow((1+rate),n)           
        total_value += face_value / math.pow((1+rate),periods)
        return total_value 
     
 #calls the ytc and ytm functions and checks for valid inputs 
    def yielding(self):
        if self.annual_coupon_rate < 0 or self.redemption <= 0 or self.purchase_price <= 0:
            print("Please Check for Valid Inputs")
            return 
        if self.settlement_date >= self.maturity_date:
            print("Not a valid Settlement Date")
            return
        
        print("------------------------------------------------")
        
    
        ytc = self.ytc()
        print ("Yield to Call" + ytc)
        ytm = self.ytm()
        print("Yield to Maturity" + ytm)
        if ytc < 0 or ytc > ytm:
            print ("Yield : " + ytm)
        else:
            print ("Yield : " + ytc)

        
#uses the present value2 function to calculate the yield to call
#with the given call date or next coupon date if it is already passed     
    def ytc(self):
        coupon = self.redemption * self.annual_coupon_rate
        ytc = self.annual_coupon_rate 
        condition = True 
        while condition:
            if (self.purchase_price < self.redemption):
                ytc += 0.00001
            else:
                ytc -= 0.00001
                
            total_pv = self.present_value2(coupon/self.payment_frequency,ytc/self.payment_frequency)
            
            if (self.purchase_price < self.redemption):
                condition = total_pv > self.purchase_price
            else:
                condition = total_pv < self.purchase_price
        print(ytc*100.0)
  
#uses the present value function to calculate the yield to maturity 
#with the given maturity date or 12/31/2099 if none 
    def ytm(self):
        coupon = self.redemption * self.annual_coupon_rate
        ytm = self.annual_coupon_rate 
        condition = True 
        while condition:
            if (self.purchase_price < self.redemption):
                ytm += 0.00001
            else:
                ytm -= 0.00001
                
            total_pv = self.present_value(coupon/self.payment_frequency,ytm/self.payment_frequency)
            
            if (self.purchase_price < self.redemption):
                condition = total_pv > self.purchase_price
            else:
                condition = total_pv < self.purchase_price
        print(ytm*100.0)
        

#reads in the file (make sure no space at top and to change the file name as necessary) 
#creates a dictionary with cusip number as key and bond object as value
def read_file():
    excel_file = r'/Users/rosechrin/Desktop/dataFile.xlsx'
    bondInfo = pd.read_excel(excel_file)
    now = str(datetime.datetime.now().year) + "/" + str(datetime.datetime.now().month) + "/" + str(datetime.datetime.now().day + 2)
    cusip_dict = {}
    for row in bondInfo.itertuples():
        bond = Bond(now, row.first_call_date, row.next_coupon_date, 
                    row.maturity_date, row.coupon, row.ask_price, row.call_maturity_price, row.payments_per_year)
        security = {row.cusip : bond}
        cusip_dict.update(security)        
    return (cusip_dict)

#iterates through the file run the ytc and ytm for every security and 
#inputs the value in the same row in an appended column 
def main():
    bond_dict = read_file()
    book = openpyxl.load_workbook('/Users/rosechrin/Desktop/dataFile.xlsx')
    sheet = book["Sheet1"]
    r = sheet.max_row
    for i in range (1,r+1):
        for k,v in bond_dict.items():
            yields = v.yielding()
            sheet.cell(row = i,column = 10).value = yields
    book.save("/Users/rosechrin/Desktop/dataFile.xlsx")


if __name__ == '__main__':
    main()